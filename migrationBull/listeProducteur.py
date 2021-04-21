# -*- coding: utf8 -*-
import openpyxl
from catTools import config
import mysql.connector
"""
fonction permettant de repucerer tout les producteur du fichier
"""


def recup_producteur(cursor):
 try:
    wb = openpyxl.load_workbook(filename='c:\\Users\\jazzt\\desktop\\NAM-IP\\bull.xlsm')
    ws = wb['Inventaire']

    for row in ws.iter_rows(min_row=config.min_row, max_col=config.max_column, max_row=config.max_row):
        prod = row[8].value
        idA = row[0].value
        id = None
        sql = "SELECT id_producteur  FROM producteurs WHERE producteur =\'" + str(prod).capitalize() + "\'"
        cursor.execute(sql)
        res = cursor.fetchone()
        if res:
           """rien ne se passe"""
        else:
            if (prod is int) or (prod is None) or (prod == 1 or prod == 0):
                config.logging.warning("artefact:"+str(idA)+";producteur n'est pas correct ou vide")
            else:
              sql1 ="INSERT INTO producteurs (producteur) VALUES(\'"+str(prod).capitalize()+"\')"
              cursor.execute(sql1)

 except mysql.connector.errors.DatabaseError as e:
    config.logging.error("artefact:" + str(id) + ";Error %d; %s" % (e.args[0], e.args[1]))


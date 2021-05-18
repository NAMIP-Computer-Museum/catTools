# -*- coding: utf8 -*-
import openpyxl
from catTools import config
"""
fonction qui va recuperer tout les usages du fichiers
mettre constante ou fichier de config 
"""


def recup_usage(cursor):
    wb = openpyxl.load_workbook(filename='c:\\Users\\jazzt\\desktop\\NAM-IP\\bull.xlsm')
    ws = wb['Inventaire']

    for row in ws.iter_rows(min_row=config.min_row, max_col=config.max_column, max_row=config.max_row):
        usage = row[1].value
        idA = row[0].value
        sql = "SELECT id_usage FROM usages WHERE libelleUsage =\'" + str(usage).upper() + "\'"
        cursor.execute(sql)
        res = cursor.fetchone()
        # condition pour voir si l'usage existe déja en DB
        if res:
           """rien ne se passe"""
        else:
            # si l'usage est vide
            if usage is None:
                config.logging.warning("artefact:"+str(idA)+";usage est vide")
            # si l'usage ne correspond pas à une valeur de base du fichier
            elif  (usage is int) or  (usage == 1 or usage == 0):
                config.logging.warning("artefact:" + str(idA) + ";usage est incorrect;"+str(usage))
            else:
              sql1 ="INSERT INTO usages (libelleUsage) VALUES(\'"+str(usage).upper()+"\')"
              cursor.execute(sql1)


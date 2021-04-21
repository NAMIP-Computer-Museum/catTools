# -*- coding: utf8 -*-
import openpyxl
from catTools import config
import mysql.connector
"""
fonction qui va recuperer tout les usages du fichiers
mettre constante ou fichier de config 
"""


def recup_conditionnement(cursor):
   try:
    wb = openpyxl.load_workbook(filename='c:\\Users\\jazzt\\desktop\\NAM-IP\\bull.xlsm')
    ws = wb['Inventaire']
    for row in ws.iter_rows(min_row=config.min_row, max_col=config.max_column, max_row=config.max_row):
        local = row[18].value
        cond = row[17].value
        idA = row[0].value
        sql = "SELECT id_cond FROM conditionnements WHERE  conditionnement =\'" + str(cond).capitalize() + "\'"
        cursor.execute(sql)
        res = cursor.fetchone()
        if res:
           """rien ne se passe"""
        else:
            if (cond is int) or (cond is None) or (cond ==1 or cond == 0):
                config.logging.warning("artefact:"+str(idA)+";conditionnement n'est pas correct ou vide")
            else:
              sqlUsage ="SELECT id_local from localisations where localisation =\'"+str(local).capitalize()+"\'"
              cursor.execute(sqlUsage)
              res = cursor.fetchall()
              for resultat in res:
                  idU = resultat[0]

              sql ="INSERT INTO conditionnements (conditionnement,localisation_key) VALUES(\'"+str(cond).capitalize()+"\',"+str(idU)+")"
              cursor.execute(sql)
   except mysql.connector.errors.DatabaseError as e:
       config.logging.error("artefact:" + str(id) + ";Error %d; %s" % (e.args[0], e.args[1]))
       config.logging.error("requete pour cond;" + str(sql) + "\n")


# -*- coding: utf8 -*-
import openpyxl
from catTools import config
"""
fonction qui va recuperer tout les usages du fichiers
mettre constante ou fichier de config 
"""


def recup_localisation(cursor):
    wb = openpyxl.load_workbook(filename='c:\\Users\\jazzt\\desktop\\NAM-IP\\bull.xlsm')
    ws = wb['Inventaire']

    for row in ws.iter_rows(min_row=config.min_row, max_col=config.max_column, max_row=config.max_row):
        local = row[18].value
        idA = row[0].value
        sql = "SELECT id_local  FROM localisations WHERE localisation =\'" + str(local).capitalize() + "\'"
        cursor.execute(sql)
        res = cursor.fetchone()
        if res:
           """rien ne se passe"""
        else:
            if (local is int) or (local is None):
                config.logging.warning("artefact:"+str(idA)+"-localisation n'est pas correcte ou vide")
            else:
              sql ="INSERT INTO localisations (localisation) VALUES(\'"+str(local).capitalize()+"\')"
              cursor.execute(sql)



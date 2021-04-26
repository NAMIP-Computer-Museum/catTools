# -*- coding: utf8 -*-
import openpyxl
from catTools import config
"""
fonction qui va recuperer toutes les familles  du fichiers
atttention certaines appartenances ne sont pas ecrite de la même façon
"""


def recup_famille(cursor):
    wb = openpyxl.load_workbook(filename='c:\\Users\\jazzt\\desktop\\NAM-IP\\bull.xlsm')
    ws = wb['Inventaire']
    i = 1
    usageCount = 13
    for row in ws.iter_rows(min_row=config.min_row, max_col=config.max_column, max_row=config.max_row):
        usage = row[1].value
        famille = row[2].value
        id = None
        idA = row[0].value
        sql = "SELECT id_famille FROM familles WHERE famille =\'" + str(famille).capitalize() + "\'"
        cursor.execute(sql)
        res = cursor.fetchone()
        if res:
           """rien ne se passe"""
        else:
            if famille is None:
                config.logging.warning("artefact:"+str(idA)+";famille est vide")
            elif (famille is int) or (famille ==1 or famille == 0):
                config.logging.warning("artefact:" + str(idA) + ";famille est incorrecte;"+str(famille))
            else:
              sqlUsage ="SELECT id_usage from usages where libelleUsage =\'"+str(usage).upper()+"\'"
              cursor.execute(sqlUsage)
              res = cursor.fetchall()
              for resultat in res:
                  idU = resultat[0]
              sql1 ="INSERT INTO familles (famille,usage_key) VALUES(\'"+str(famille).capitalize()+"\',"+str(idU)+")"
              cursor.execute(sql1)

    while i <= usageCount:
        sql ="INSERT INTO familles (famille,usage_key) VALUES(\'inconnue\',"+str(i)+")"
        cursor.execute(sql)
        i = i+1


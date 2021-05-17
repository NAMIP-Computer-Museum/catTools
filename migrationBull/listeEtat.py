# -*- coding: utf8 -*-

import openpyxl
from catTools import config
"""
fonction permettant de repucerer tout les etats du fichier
"""


def recup_etat(cursor):
    wb = openpyxl.load_workbook(filename='c:\\Users\\jazzt\\desktop\\NAM-IP\\bull.xlsm')
    ws = wb['Inventaire']

    for row in ws.iter_rows(min_row=config.min_row, max_col=config.max_column, max_row=config.max_row):
        etat = row[11].value
        idA = row[0].value
        #verification si l'etat exsite dans la DB
        sql = "SELECT id_etat FROM etats WHERE etat =\'" + str(etat).upper() + "\'"
        cursor.execute(sql)
        res = cursor.fetchone()
        if res:
           """rien ne se passe"""
        else:
            if (etat is None):
                config.logging.warning("artefact:"+str(idA)+";etat est vide")
            elif (etat is int) or (etat == 1 or etat == 0):
                """config.logging.warning("artefact:" + str(idA) + ";etat est incorrecte;"+str(etat))"""
            else:
                sql1 = "INSERT INTO etats (etat) VALUES(\'" + str(etat).upper() + "\')"
                cursor.execute(sql1)



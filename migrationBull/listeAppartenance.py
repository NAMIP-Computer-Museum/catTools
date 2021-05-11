# -*- coding: utf8 -*-
import openpyxl
from catTools import config
"""
fonction permettant de repucerer toutes les appartenances du fichier
atttention certaines appartenances ne sont pas ecrite de la même façon
"""


def recup_appartenance(cursor):
    wb = openpyxl.load_workbook(filename='c:\\Users\\jazzt\\desktop\\NAM-IP\\bull.xlsm')
    ws = wb['Inventaire']

    for row in ws.iter_rows(min_row=config.min_row, max_col=config.max_column, max_row=config.max_row):
        appart = row[5].value
        idA = row[0].value
        sql = "SELECT id_appart FROM appartenances WHERE appartenance =\'" + str(appart)+ "\'"
        cursor.execute(sql)
        res = cursor.fetchone()
        if res:
           """rien ne se passe"""
        else:
            if (appart is None) or (appart == 0) :
                config.logging.warning("artefact:"+str(idA)+";appartenance est vide")
            elif   (appart is int) or  (appart ==1):
                config.logging.warning("artefact:" + str(idA) + ";appartenance est incorrecte;"+str(appart))
            else:
              sql1 ="INSERT INTO appartenances (appartenance) VALUES(\'"+str(appart).capitalize()+"\')"
              cursor.execute(sql1)

# -*- coding: utf8 -*-
import openpyxl
import re
from catTools import config
import datetime
import mysql.connector
import listeAppartenance
import listeConditionnement
import listeEtat
import listeUsage
import listeFamille
import listeProducteur
import listeLocalisation
"""
methode pour savoir si id est rempli 
"""


def process_id(row):
    # print(row[0].value)
    id = row[0].value
    return id


"""
methode pour savoir si usage est present
doit être entre 2 et 3 lettres majuscule  ^[A-Z]{2,3}$
"""


def process_usage(row,conn):
    # print(row[1].value)
    usage = row[1].value
    p = re.compile('^[A-Z]{2,3}$')
    if usage is None:
        config.logging.warning("usage non encodé")
        return None
    elif p.match(usage):

        cursor = conn.cursor()
        sql = "SELECT id_usage FROM usages WHERE libellUsage =\'" + str(usage).upper() + "\'"
        cursor.execute(sql)
        res = cursor.fetchall()
        for resultat in res:
            id = resultat[0]
        return id
    else:
        config.logging.warning("usage doit être compris en 2 et 3 lettres en majuscule")
        return None


"""
methode pour voir si la famille est présente ou non
"""


def process_famille(row,conn):
    # print(row[2].value)
    famille = row[2].value
    idA = row[0].value
    id = None
    if famille is None:
        config.logging.warning("artefact:"+str(idA)+";famille vide")
        return None
    else:
        cursor = conn.cursor()
        sql = "SELECT id_famille FROM familles WHERE famille =\'" + str(famille).capitalize() + "\'"
        cursor.execute(sql)
        res = cursor.fetchall()
        for resultat in res:
            id = resultat[0]
    return id


"""
methode pour voir si le libellé est présente
"""


def process_libelle(row):
    # print(row[3].value)
    libelle = row[3].value
    if str(libelle).__contains__("\'"):
     # escape des caracères ' et "
     libelle = re.sub('[\']', '', libelle)

    return libelle


"""
  methode pour le modele de l'objet
"""


def process_modele(row):
    modele = row[4].value
    if modele is None:
        return None
    else:
        return modele


"""
methode pour l'apparetenance est présente
"""


def process_appartenance(row,conn):
    appart = row[5].value
    id = None
    if appart is None:
        return None
    else:
        cursor = conn.cursor()
        sql = "SELECT id_appart FROM appartenances WHERE appartenance =\'" + str(appart).capitalize() + "\'"
        cursor.execute(sql)
        res = cursor.fetchall()
        for resultat in res:
            id = resultat[0]
    return id


"""
methode pour n° de serie ou numero de la machine
"""


def process_numero_serie(row):
    nums = row[6].value
    if nums is None:
        return None
    else:
        return nums


"""
methode pour année de production
en 4 chiffres 
"""


def process_anprod(row):
    anprod = row[7].value
    if anprod is None:
        return None
    else:
        return anprod


"""
methode pour le producteur
un seul producteur 
"""


def process_producteur(row,conn):
    prod = row[8].value
    id = None
    if prod is None:
        prod = None
    elif prod == 0:
        prod = None
    else:
        if str(prod).__contains__("\'"):
         # escape des caracères ' et "
         prod = re.sub('[\'"]', '', prod)

    cursor = conn.cursor()
    sql = "SELECT id_producteur FROM producteurs WHERE producteur =\'" + str(prod).capitalize() + "\'"
    cursor.execute(sql)
    res = cursor.fetchall()
    for resultat in res:
        id = resultat[0]
    return id


"""
methode pour la quantité
peut etre compris entre 0 et autant que possible
"""


def process_quantite(row):
    # print(row[9].value)
    qte = row[9].value
    id = row[0].value
    if qte is None:
        config.logging.warning("artefact:"+str(id)+";pas de quantité")
        return None
    else:
        return qte


"""
methode pour l'etat de l'artefact
peut etre de différents état
"""


def process_etat(row,conn):
    # print(row[5].value)
    etat = row[5].value
    if etat is None:
        return None
    else:
        cursor = conn.cursor()
        sql = "SELECT id_etat FROM etats WHERE etat =\'" + str(etat).upper() + "\'"
        cursor.execute(sql)
        res = cursor.fetchall()
        id = None
        for resultat in res:
            id = resultat[0]
    return id


"""
methode pour la conditionnement ok
pas obligatoire donc verifier si il y en a un 
"""


def process_conditionement(row,conn):
    # print(row[17].value)
    cond = row[17].value
    id = None
    if cond is None:
        cond = None
    elif cond == 0:
        cond = None
    else:
        if str(cond).__contains__("\'"):
         # escape des caracères ' et "
         cond = re.sub('[\'"]', '', cond)

    cursor = conn.cursor()
    sql = "SELECT id_cond FROM conditionnements WHERE conditionnement = \'" + str(cond).capitalize() + "\'"
    cursor.execute(sql)
    res = cursor.fetchall()
    for resultat in res:
        id = resultat[0]
    return id


"""
methode pour la localisation ok
^C[1-9]-[A-Z][1-20]?-[1-20]$  => container
^B-[A-Z][1-20].[1-20]$ => container documents/revues
^Ate.[A-Z][1-9]-[1-9]$ => atelier
NT => non_trouve
vitrine
S => sol
M => mur
table clucth
etabli ou etabli mobile
musée ou musée B 
"""


def process_localisation(row,conn):
    local = row[18].value
    id = None
    cursor = conn.cursor()
    sql = "SELECT id_local FROM localisations WHERE localisation = \'" + str(local).capitalize() + "\'"
    cursor.execute(sql)
    res = cursor.fetchall()
    for resultat in res:
        id = resultat[0]
    return id



"""
methode pour la longueur
nombre+cm  ^[A-Z]{2,3}$
"""


def process_longueur(row):
    longueur = row[19].value
    id = row[0].value
    cm = re.compile('^[1-9]+Cm$')
    if longueur is None:
        config.logging.warning("artefact:"+str(id)+";pas de longueur")
        return None
    elif longueur is int:
        return longueur
    elif cm.match(str(longueur)):
        str(longueur).split("C")
        return longueur[0]
    else:
         config.logging.warning("artefact:"+str(id)+";mauvais encodage de la longueur")
    return None


"""
methode pour la largeur
nombre+cm ^[A-Z]{2,3}$
"""


def process_largeur(row):
    # print(row[20].value)
    largeur = row[20].value
    id = row[0].value
    cm = re.compile('^[1-9]+Cm$')
    if largeur is None:
        config.logging.warning("artefact:"+str(id)+";pas de largeur")
        return None
    elif largeur is int:
        return largeur
    elif cm.match(str(largeur)):
        str(largeur).split("C")
        config.logging.debug(largeur[0])
        return largeur[0]
    else:
        config.logging.warning("artefact:"+str(id)+";mauvais encodage de la largeur")
        return None

"""
methode pour la hauteur
nombre+cm  ^[A-Z]{2,3}$

"""


def process_hauteur(row):
    hauteur = row[21].value
    id = row[0].value
    cm = re.compile('^[1-9]+Cm$')
    if hauteur is None:
        config.logging.warning("artefact:"+str(id)+";pas de hauteur")
        return None
    elif hauteur is int:
        return hauteur
    elif cm.match(str(hauteur)):
        str(hauteur).split("C")
        return hauteur[0]
    else:
        config.logging.warning("artefact:"+str(id)+";mauvais encodage de la hauteur")
        return None

"""
methode pour le poids
en kg
nombre+kg  ^[A-Z]{2,3}$
attention peuty avoir des gramme gr
"""


def process_poids(row):
    poids = row[22].value
    id = row[0].value
    kg = re.compile('^[1-9]+Kg$')
    if poids is None:
        config.logging.warning("artefact:"+str(id)+";pas de poids")
        return None
    elif poids is int:
        return poids
    elif kg.match(str(poids)):
        str(poids).split("K")
        return poids[0]
    else:
       config.logging.warning("artefact:"+str(id)+";mauvais encodage du poids")
    return None


"""
methode pour le donateur
pas obligatoire 
"""


def process_donateur(row,conn):
    donateur = row[26].value
    id = None
    cursor = conn.cursor()
    sql = "SELECT id_donateur FROM donateurs WHERE donateur =\'" + str(donateur) + "\'"
    cursor.execute(sql)
    res = cursor.fetchall()
    for resultat in res:
        id = resultat[0]
    if id is not None:
        return id
    else:
        return donateur


"""
methode pour les liens
pas obligatoire
"""


def process_liens(row):
    liens = (row[27].value, row[28].value)
    return liens


"""
methode pour commentaire
pas obligatoire
"""


def process_comment(row):
    comment = row[29].value
    if comment is None:
        comment = None
    elif comment == 0:
        comment = None
    else:
        if str(comment).__contains__("\'"):
         # escape des caracères ' et "
         comment = re.sub('[\'"]', '', comment)
    return comment


"""
methode pour les images
pas obligatoire
mettre le chemin d'acces
"""


def process_images(row):
    return row[30].value


"""
methode pour la date d'entrée dans le musée
obligatoire rentré en dd-mm-aaaa
cherche encore comment verifier la date car il faut du String pas du datetime
"""


def process_datein(row):
    datetime.date = row[31].value
    id = row[0].value
    dateIn = None
    if datetime.date is None:
        dateIn = None
        config.logging.warning("artefact:" + str(id) + ";pas de date d'entrée")
    else:
        dateIn = datetime.date
    return dateIn


"""
methode pour la date de recolement
pas obligatoire mais date 1er recolement = date in
dd-mm-aaaa
"""


def process_recolement(row):
    datetime.date = row[33].value
    id = row[0].value
    if datetime.date is None:
        return None
        config.logging.warning("artefact:" + str(id) + ";pas de recolement")
    else:
        return datetime.date


"""
methode pour voir si la ligne est remplie correctement
retournera les retour des autres fonctions pour montrer ce qui est à changer ou 
à remplir
"""


def process_row(row,conn):

     id = process_id(row)
     famille = process_famille(row,conn)
     libelle = process_libelle(row)
     modele = process_modele(row)
     appart = process_appartenance(row,conn)
     numSerie = process_numero_serie(row)
     anProd = process_anprod(row)
     prod = process_producteur(row,conn)
     qte = process_quantite(row)
     etat = process_etat(row,conn)
     cond = process_conditionement(row,conn)
     local = process_localisation(row,conn)
     long = process_longueur(row)
     larg = process_largeur(row)
     haut = process_hauteur(row)
     poids = process_poids(row)
     donateur = process_donateur(row,conn)
     lien = process_liens(row)
     comment = process_comment(row)
     image = process_images(row)
     dateIn = process_datein(row)
     recolement = process_recolement(row)

     cursor = conn.cursor()
     if donateur == 0:
         idDon = None
     elif isinstance(donateur, int):
         idDon = donateur

     else:
         # insertion dans la table de donateur
         sqlDon = "INSERT INTO donateurs (donateur) VALUES (\'" + str(donateur) + "\')"
         try:
             cursor.execute(sqlDon)
         except mysql.connector.errors.DatabaseError as e:
             config.logging.error("artefact:" + str(id) + ";Error %d; %s;"+str(sqlDon)+"\n" % (e.args[0], e.args[1]))
             conn.rollback()

         # recuperation de l'id du donateur
         sqlDon2 = "SELECT id_donateur FROM donateurs WHERE donateur = \'" + str(donateur) + "\'"
         try:
            cursor.execute(sqlDon2)
         except mysql.connector.errors.DatabaseError as e:
            config.logging.error("artefact:" + str(id) + ";Error %d; %s;"+ str(sqlDon2)+"\n" % (e.args[0], e.args[1]))
            conn.rollback()
         res = cursor.fetchall()
         for resultat in res:
           idDon = resultat[0]

     p1 = "INSERT INTO artefacts (`id_artefact`, `libelle`, `modele`, `numSerie`, `anProd`, `quantite`, `dateIn`, `longueur`,"
     p2 = "`largeur`, `hauteur`, `poids`, `commentaire`, `donateur_key`, `cond_key`, `prod_key`, `etat_key`, `localisation_key`,"
     p3 = "`appart_key`, `famille_key`) VALUES(" + str(id) + ",\'" + str(libelle) + "\',\'" + str(modele) + "\',\'" + str(numSerie)
     if dateIn is not None:
        p4 = "\'," + str(anProd) + "," + str(qte) + ",\'" + str(dateIn) + "\'," + str(long)+ "," +str(larg)+ "," +str(haut)+ "," +str(poids)
     else:
        p4 = "\'," + str(anProd) + "," + str(qte) + "," + str(dateIn) + "," +str(long)+ "," +str(larg)+ "," +str(haut)+ "," + str(poids)
     p5 = ",\'" + str(comment) + "\'," + str(idDon) + "," + str(cond) + "," + str(prod) + "," + str(etat) + "," + str(local)
     p6 = "," + str(appart) + "," + str(famille) + ")"
     sqlArtefact = p1 + p2 + p3 + p4 + p5 + p6
     sqlArtefact = re.sub('None', 'NULL', sqlArtefact)
     try:
      cursor.execute(sqlArtefact)
     except mysql.connector.errors.DatabaseError as e:
         config.logging.error("artefact:" + str(id) + ";Error %d; %s" % (e.args[0], e.args[1]))
         config.logging.error("artefact:"+str(id)+";; erreur requête;"+str(sqlArtefact)+"\n")
         conn.rollback()
    # insertion dans la table des liens
     """ 
     for l in lien:
       if l is None:
         continue
       elif l == 0:
        continue
       else:
         sqllien = "INSERT INTO `liens`(`principale`, `secondaire`) VALUES ("+str(id)+","+str(l)+")"
         print(sqllien)
         cursor.execute(sqllien)
     """
     # insertion dans la table images
     sqlImage = "INSERT INTO `images`(`image`, `artefact_key`) VALUES (\'" + str(image) + "\'," + str(id) + ")"
     cursor.execute(sqlImage)

    # insertion dans la table recolements
     if recolement is None:
         #print("")
         ""
     else:
        sqlRecol = "INSERT INTO `recolements`(`recolement`, `artefact_key`) VALUES (\'" + str(recolement) + "\'," + str(id) + ")"
        try:
         cursor.execute(sqlRecol)
        except mysql.connector.errors.DatabaseError as e:
            config.logging.error("artefact:" + str(id) + ";Error %d;%s;"+str(sqlRecol)+"\n" % (e.args[0], e.args[1]))
            conn.rollback()
     conn.commit()

wb = openpyxl.load_workbook(filename='c:\\Users\\jazzt\\desktop\\NAM-IP\\bull.xlsm')
ws = wb['Inventaire']
"""
boucle pour vérifier si les cellules d'une ligne sont compléter ou non et 
si oui verifier si elle sont bien encodée + mise dans la DataBase si tout est OK 
"""
try:
 conn = mysql.connector.connect(host=config.host, user=config.user, password=config.passwd, database=config.database)
 cursor = conn.cursor()
 listeEtat.recup_etat(cursor)
 listeProducteur.recup_producteur(cursor)
 listeUsage.recup_usage(cursor)
 listeAppartenance.recup_appartenance(cursor)
 listeFamille.recup_famille(cursor)
 listeLocalisation.recup_localisation(cursor)
 listeConditionnement.recup_conditionnement(cursor)
 conn.commit()
 for row in ws.iter_rows(min_row=config.min_row, max_col=config.max_column, max_row=config.max_row):
    if all([cell.value is None for cell in row[2:]]):
        continue
    process_row(row,conn)
finally:
    conn.close
from catTools import config
import sqlExport
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
def stylesheet(ws):
    # colonne des id artéfact
    ws.column_dimensions['A'].width = 25
    ws["A1"] = "id"
    ws["A1"].font = config.font
    # colonne des fichiers sources
    ws.column_dimensions['B'].width = 23
    ws["B1"] = "source file"
    ws["B1"].font = config.font
    # colonne des libelles
    ws.column_dimensions['C'].width = 35
    ws["C1"] = "libellé"
    ws["C1"].font = config.font
    # colonne des numSerie
    ws.column_dimensions['D'].width = 15
    ws["D1"] = "numéro série"
    ws["D1"].font = config.font
    # colonne des anprod
    ws.column_dimensions['E'].width = 19
    ws["E1"] = "année production"
    ws["E1"].font = config.font
    # colonne des dates in
    ws.column_dimensions['F'].width = 15
    ws["F1"] = "date entrée"
    ws["F1"].font = config.font
    # colonne des longueurs
    ws["G1"] = "longueur"
    ws["G1"].font = config.font
    # colonne des largeurs
    ws["H1"] = "largeur"
    ws["H1"].font = config.font
    # colonne des hauteurs
    ws["I1"] = "hauteur"
    ws["I1"].font = config.font
    # colonne des poids
    ws["J1"] = "poids"
    ws["J1"].font = config.font
    # colonne des donateurs
    ws["K1"] = "donateur"
    ws["K1"].font = config.font
    # colonne des producteurs
    ws["L1"] = "producteur"
    ws["L1"].font = config.font
    # colonne des etats
    ws["M1"] = "état"
    ws["M1"].font = config.font
    # colonne des localisations
    ws["N1"] = "localisation"
    ws["N1"].font = config.font
def insertArtefact(ws,cursor):
    # dico d'info à mettre dans l'excel pour une card
    card = {}
    cursor.execute(sqlExport.sqlAr)
    resultat = cursor.fetchall()
    for res in resultat:
        """changer valeur null par vide """
        card['A'] = res[0]
        card['B'] = res[1]
        card['C'] = res[2]
        card['D'] = res[3]
        card['E'] = res[4]
        card['F'] = res[5]
        card['G'] = res[6]
        card['H'] = res[7]
        card['I'] = res[8]
        card['J'] = res[9]
        if res[10] is not None:
         sql = str(sqlExport.sqlD)+str(res[10])
         cursor.execute(sql)
         resultat = cursor.fetchone()
         card['K'] = resultat[0]
        if res[11] is not None:
         sql = str(sqlExport.sqlP)+str(res[11])
         cursor.execute(sql)
         resultat = cursor.fetchone()
         card['L'] = resultat[0]
        if res[12] is not None:
         sql = str(sqlExport.sqlE)+str(res[12])
         cursor.execute(sql)
         resultat = cursor.fetchone()
         card['M'] = resultat[0]
        if res[13] is not None:
         sql = str(sqlExport.sqlLocal)+str(res[13])
         cursor.execute(sql)
         resultat = cursor.fetchone()
         card['N'] = resultat[0]
        ws.append(card)
        card.clear()